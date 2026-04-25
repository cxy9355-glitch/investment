from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"G:\Codex\个人\investment")
BASE_WORKBOOK = ROOT / "机构持仓研究" / "长期持有全池数据集_2026-04-17_可比清理版.xlsx"
DISPLAY_WORKBOOK = ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_长期持有审美分类总表_2026-04-17.xlsx"
RESEARCH_WORKBOOK = ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_研究总表_2026-04-17.xlsx"
SCREEN_WORKBOOK = ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_历史持仓筛选_2026-04-17_可比清理版.xlsx"

SNAPSHOT_DATE = "2026-04-17"
CHECKED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

STANDARD_SHEETS = [
    "company_master",
    "holding_timeline",
    "operating_timeline",
    "classification_snapshot",
    "research_judgement",
    "tag_dictionary",
    "source_checks",
]

REQUIRED_PANEL_FIELDS = [
    "代码",
    "中文公司名",
    "英文公司名",
    "市场",
    "当前持有机构",
    "当前持有状态",
    "覆盖机构数",
    "出现季度数",
    "覆盖年份数",
    "当前持有机构数",
    "平均持仓权重",
    "峰值持仓权重",
    "有效财年行数",
    "核心完整财年行数",
    "5年平均ROE",
    "5年平均ROA",
    "持有持续性分数",
    "持有持续性分位",
    "经营持续性分数",
    "经营持续性分位",
    "分类结果",
    "观察原因",
]


def company_id(market: object, ticker: object) -> str:
    market_text = "" if pd.isna(market) else str(market).strip()
    ticker_text = "" if pd.isna(ticker) else str(ticker).strip()
    return f"{market_text}|{ticker_text}"


def norm_name(value: object) -> str:
    if pd.isna(value):
        return ""
    return "".join(ch for ch in str(value).upper() if ch.isalnum())


def add_company_id(df: pd.DataFrame, ticker_col: str = "代码", market_col: str = "市场") -> pd.DataFrame:
    work = df.copy()
    if ticker_col not in work.columns and "ticker" in work.columns:
        ticker_col = "ticker"
    if market_col not in work.columns:
        work[market_col] = ""
    work.insert(0, "company_id", [company_id(m, t) for m, t in zip(work[market_col], work[ticker_col])])
    return work


def build_company_master(coverage: pd.DataFrame, classification: pd.DataFrame) -> pd.DataFrame:
    left = coverage.copy()
    right = classification.copy()
    if "代码" not in left.columns and "ticker" in left.columns:
        left = left.rename(columns={"ticker": "代码"})

    base_cols = ["代码", "中文公司名", "英文公司名", "市场"]
    quality_cols = ["有效财年行数", "核心完整财年行数", "是否满足主池完整度"]
    merged = pd.concat(
        [
            left[[c for c in base_cols + quality_cols + ["当前是否仍持有", "当前持有机构"] if c in left.columns]],
            right[[c for c in base_cols + quality_cols + ["当前持有状态", "当前持有机构", "分类结果", "观察原因"] if c in right.columns]],
        ],
        ignore_index=True,
    )

    classified = add_company_id(right[[c for c in base_cols + ["分类结果", "观察原因"] if c in right.columns]])
    classified_lookup: dict[str, pd.Series] = {}
    for _, item in classified.iterrows():
        for name_col in ["英文公司名", "中文公司名"]:
            key = norm_name(item.get(name_col))
            if key:
                classified_lookup.setdefault(key, item)

    unresolved = merged["代码"].isna() | merged.get("市场", pd.Series(index=merged.index, dtype=object)).eq("映射待核")
    for idx, row in merged[unresolved].iterrows():
        key = norm_name(row.get("英文公司名")) or norm_name(row.get("中文公司名"))
        match = classified_lookup.get(key)
        if match is None:
            continue
        for col in base_cols + ["分类结果", "观察原因"]:
            if col in match.index:
                merged.at[idx, col] = match[col]

    merged = add_company_id(merged)
    merged = merged.sort_values(["company_id"]).drop_duplicates("company_id", keep="last")

    columns = [
        "company_id",
        "代码",
        "中文公司名",
        "英文公司名",
        "市场",
        "当前是否仍持有",
        "当前持有状态",
        "当前持有机构",
        "有效财年行数",
        "核心完整财年行数",
        "是否满足主池完整度",
        "是否进入分类快照",
        "分类口径状态",
        "未进入分类原因",
        "分类结果",
        "观察原因",
    ]
    for col in columns:
        if col not in merged.columns:
            merged[col] = pd.NA
    blank_class = merged["分类结果"].isna() | merged["分类结果"].astype(str).str.strip().eq("")
    merged.loc[blank_class, "分类结果"] = "未进入分类样本"
    merged["是否进入分类快照"] = ~blank_class
    merged["分类口径状态"] = merged["是否进入分类快照"].map({True: "已进入四象限分类", False: "未进入四象限分类"})
    merged["未进入分类原因"] = pd.NA
    insufficient = blank_class & (
        merged["是否满足主池完整度"].eq(False)
        | pd.to_numeric(merged["有效财年行数"], errors="coerce").fillna(0).lt(5)
    )
    merged.loc[blank_class, "未进入分类原因"] = "保留在全池主数据中，但未进入当前四象限分类口径"
    merged.loc[insufficient, "未进入分类原因"] = "有效财年或核心字段覆盖不足，暂未进入四象限分类"
    merged.loc[blank_class, "观察原因"] = merged.loc[blank_class, "未进入分类原因"]
    return merged[columns].reset_index(drop=True)


def build_research_judgement(company_master: pd.DataFrame) -> pd.DataFrame:
    rows = company_master[["company_id", "代码", "分类结果"]].copy()
    rows = rows.rename(columns={"代码": "ticker", "分类结果": "case_type"})
    rows.insert(2, "manager_scope", "全机构合并")
    rows.insert(3, "judgement_version", "v1")
    for col in [
        "entry_tags",
        "entry_summary",
        "entry_confidence",
        "hold_tags",
        "hold_summary",
        "hold_confidence",
        "exit_tags",
        "exit_summary",
        "exit_confidence",
        "degradation_tags",
        "degradation_summary",
        "degradation_confidence",
        "degradation_nature",
        "exit_hindsight",
        "evidence_refs",
        "last_reviewed_at",
    ]:
        rows[col] = pd.NA
    ordered = [
        "company_id",
        "ticker",
        "manager_scope",
        "judgement_version",
        "entry_tags",
        "entry_summary",
        "entry_confidence",
        "hold_tags",
        "hold_summary",
        "hold_confidence",
        "exit_tags",
        "exit_summary",
        "exit_confidence",
        "degradation_tags",
        "degradation_summary",
        "degradation_confidence",
        "degradation_nature",
        "exit_hindsight",
        "case_type",
        "evidence_refs",
        "last_reviewed_at",
    ]
    return rows[ordered]


def build_tag_dictionary() -> pd.DataFrame:
    return pd.DataFrame(columns=["tag_name", "tag_group", "definition", "status", "notes"])


def build_source_checks(
    coverage: pd.DataFrame,
    holdings: pd.DataFrame,
    annual: pd.DataFrame,
    classification: pd.DataFrame,
    company_master: pd.DataFrame,
) -> pd.DataFrame:
    checks: list[dict[str, object]] = []

    def add(group: str, workbook: Path, sheet: str, field: str, status: str, notes: str) -> None:
        checks.append(
            {
                "check_group": group,
                "workbook": workbook.name,
                "sheet": sheet,
                "field": field,
                "status": status,
                "notes": notes,
                "checked_at": CHECKED_AT,
            }
        )

    for sheet, frame, required in [
        ("可比池覆盖说明", coverage, ["代码", "中文公司名", "英文公司名", "市场"]),
        ("实体级持仓历史_可比池", holdings, ["ticker", "manager_entity", "report_date", "position_weight"]),
        ("年度经营质量_可比池", annual, ["ticker", "fiscal_year", "roe", "roa", "核心字段完整"]),
    ]:
        for field in required:
            add(
                "base_source",
                BASE_WORKBOOK,
                sheet,
                field,
                "ok" if field in frame.columns else "missing",
                "标准底表来源字段",
            )

    missing = [field for field in REQUIRED_PANEL_FIELDS if field not in classification.columns]
    add(
        "panel_contract",
        DISPLAY_WORKBOOK,
        "分类总表",
        ",".join(REQUIRED_PANEL_FIELDS),
        "ok" if not missing else "missing",
        "四象限面板稳定读取接口；缺失字段：" + ("无" if not missing else ",".join(missing)),
    )
    add(
        "workbook_role",
        BASE_WORKBOOK,
        ",".join(STANDARD_SHEETS),
        "standard_sheets",
        "ok",
        "标准底表 sheet 已生成；旧中文 sheet 暂保留作兼容来源。",
    )
    add(
        "workbook_role",
        RESEARCH_WORKBOOK,
        "研究主表/观察与负样本",
        "role",
        "derived",
        "辅助阅读工作簿，后续不作为新增事实字段的写入目标。",
    )
    add(
        "workbook_role",
        SCREEN_WORKBOOK,
        "历史热度原榜/估值总表等",
        "role",
        "derived",
        "历史筛选展示工作簿，后续不作为新增事实字段的写入目标。",
    )

    add(
        "data_quality",
        BASE_WORKBOOK,
        "company_master",
        "company_id",
        "ok" if company_master["company_id"].is_unique else "duplicate",
        f"company_master 行数 {len(company_master)}；company_id 唯一性检查。",
    )
    add(
        "data_quality",
        BASE_WORKBOOK,
        "company_master",
        "代码/市场/分类结果",
        "ok"
        if (
            company_master["代码"].notna().all()
            and not company_master["市场"].astype(str).eq("映射待核").any()
            and company_master["分类结果"].notna().all()
        )
        else "review",
        "检查空代码、映射待核和空分类。",
    )
    not_classified = company_master["分类口径状态"].eq("未进入四象限分类").sum()
    add(
        "data_quality",
        BASE_WORKBOOK,
        "company_master",
        "分类口径状态",
        "ok",
        f"已进入四象限分类 {len(company_master) - not_classified}；未进入四象限分类 {not_classified}。",
    )
    return pd.DataFrame(checks)


def write_sheets(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)


def style_standard_sheets(path: Path, sheet_names: list[str]) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            sample = [str(cell.value or "") for cell in column_cells[: min(ws.max_row, 120)]]
            width = min(max(len(v) for v in sample) + 2, 48)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)


def main() -> None:
    coverage = pd.read_excel(BASE_WORKBOOK, sheet_name="可比池覆盖说明")
    holdings = pd.read_excel(BASE_WORKBOOK, sheet_name="实体级持仓历史_可比池")
    annual = pd.read_excel(BASE_WORKBOOK, sheet_name="年度经营质量_可比池")
    classification = pd.read_excel(DISPLAY_WORKBOOK, sheet_name="分类总表")

    company_master = build_company_master(coverage, classification)
    holding_timeline = add_company_id(holdings, ticker_col="ticker", market_col="市场")
    operating_timeline = add_company_id(annual, ticker_col="ticker", market_col="市场")
    classification_snapshot = add_company_id(classification)
    classification_snapshot.insert(1, "snapshot_date", SNAPSHOT_DATE)

    sheets = {
        "company_master": company_master,
        "holding_timeline": holding_timeline,
        "operating_timeline": operating_timeline,
        "classification_snapshot": classification_snapshot,
        "research_judgement": build_research_judgement(company_master),
        "tag_dictionary": build_tag_dictionary(),
        "source_checks": build_source_checks(coverage, holdings, annual, classification, company_master),
    }
    write_sheets(BASE_WORKBOOK, sheets)
    style_standard_sheets(BASE_WORKBOOK, list(sheets))
    print(f"updated base workbook: {BASE_WORKBOOK}")
    for name, frame in sheets.items():
        print(f"{name}: {len(frame)} rows")


if __name__ == "__main__":
    main()
