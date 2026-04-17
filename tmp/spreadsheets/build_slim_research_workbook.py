from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"G:\Codex\个人\investment")
SOURCE = ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_历史持仓筛选_2026-04-17_可比清理版.xlsx"
OUTPUT = ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_研究总表_2026-04-17.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)

PCT_COLS = {
    "ROE",
    "ROA",
    "股息率",
    "毛利率",
    "净利率",
    "收入增速",
    "利润增速",
    "平均持仓权重",
    "峰值持仓权重",
}
MONEY_COLS = {
    "累计持仓金额(亿美元)",
    "平均单季持仓金额(亿美元)",
    "单季最高持仓金额(亿美元)",
    "最新持仓金额(亿美元)",
    "当前价格",
    "当前市值(亿美元)",
}
RATIO_COLS = {"PE(TTM)", "PB", "Debt/Equity"}
DATE_COLS = {"首次出现", "最近出现", "最新财报期"}


def load_frames() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    hot = pd.read_excel(SOURCE, sheet_name="历史热度原榜")
    val = pd.read_excel(SOURCE, sheet_name="估值总表")
    caliber = pd.read_excel(SOURCE, sheet_name="口径说明")
    decisions = pd.read_excel(SOURCE, sheet_name="缺数样本判定")
    shortlist = pd.read_excel(SOURCE, sheet_name="深研候选")
    keep_missing = pd.read_excel(SOURCE, sheet_name="保留缺数案例")
    limited = pd.read_excel(SOURCE, sheet_name="有限财年覆盖样本")
    return hot, val, caliber, decisions, shortlist, keep_missing, limited


def build_main_sheet(val: pd.DataFrame, decisions: pd.DataFrame) -> pd.DataFrame:
    completeness = decisions[["代码", "有效财年行数", "核心完整财年行数", "是否满足主池完整度"]].copy()
    main = val.merge(completeness, on="代码", how="left")
    cols = [
        "代码",
        "中文公司名",
        "英文公司名",
        "市场",
        "当前持有机构",
        "覆盖机构数",
        "出现季度数",
        "最近出现",
        "PE(TTM)",
        "PB",
        "ROE",
        "ROA",
        "股息率",
        "毛利率",
        "净利率",
        "Debt/Equity",
        "收入增速",
        "利润增速",
        "当前价格",
        "当前市值(亿美元)",
        "最新财报期",
        "有效财年行数",
        "核心完整财年行数",
        "是否满足主池完整度",
    ]
    return main[cols].copy()


def build_preference_sheet(hot: pd.DataFrame) -> pd.DataFrame:
    pref = hot.copy()
    pref = pref.rename(columns={"热度排名": "历史偏好排名"})
    pref["当前持有状态"] = pref["当前持有机构"].apply(lambda x: "当前在仓" if pd.notna(x) and str(x).strip() else "当前不在仓")
    cols = [
        "历史偏好排名",
        "代码",
        "中文公司名",
        "英文公司名",
        "市场",
        "机构口径",
        "覆盖机构数",
        "首次出现",
        "最近出现",
        "出现季度数",
        "覆盖年份数",
        "累计持仓金额(亿美元)",
        "平均单季持仓金额(亿美元)",
        "单季最高持仓金额(亿美元)",
        "当前持有状态",
        "当前持有机构",
        "最新持仓金额(亿美元)",
        "备注",
    ]
    return pref[cols].copy()


def build_valuation_aux(val: pd.DataFrame) -> pd.DataFrame:
    aux = val.copy()
    aux["估值口径备注"] = "历史百分位为自算年表分位，仅作辅助参考"
    cols = [
        "代码",
        "中文公司名",
        "市场",
        "PE(TTM)",
        "PE年末历史百分位(自算)",
        "PB",
        "PB年末历史百分位(自算)",
        "综合估值百分位(自算)",
        "估值口径备注",
    ]
    return aux[cols].copy()


def build_shortlist_sheet(shortlist: pd.DataFrame, decisions: pd.DataFrame) -> pd.DataFrame:
    completeness = decisions[["代码", "有效财年行数", "核心完整财年行数", "是否满足主池完整度"]].copy()
    short = shortlist.merge(completeness, on="代码", how="left")
    cols = [
        "代码",
        "中文公司名",
        "英文公司名",
        "市场",
        "当前持有机构",
        "覆盖机构数",
        "出现季度数",
        "最近出现",
        "PE(TTM)",
        "PB",
        "ROE",
        "ROA",
        "股息率",
        "有效财年行数",
        "核心完整财年行数",
        "是否满足主池完整度",
        "候选理由",
        "关注点",
    ]
    return short[cols].copy()


def build_negative_samples_sheet(keep_missing: pd.DataFrame, limited: pd.DataFrame) -> pd.DataFrame:
    keep = keep_missing.copy()
    keep["样本类型"] = "无有效财年但保留案例"
    lim = limited.copy()
    lim["样本类型"] = "有效财年不足观察样本"
    merged = pd.concat([keep, lim], ignore_index=True, sort=False)
    cols = [
        "样本类型",
        "代码",
        "中文公司名",
        "英文公司名",
        "市场",
        "当前持有机构",
        "覆盖机构数",
        "出现季度数",
        "最近出现",
        "平均持仓权重",
        "峰值持仓权重",
        "有效财年行数",
        "核心完整财年行数",
        "判定结果",
        "判定原因",
    ]
    existing = [c for c in cols if c in merged.columns]
    out = merged[existing].copy()
    out = out.sort_values(["样本类型", "覆盖机构数", "出现季度数"], ascending=[True, False, False]).reset_index(drop=True)
    return out


def build_caliber_sheet(caliber: pd.DataFrame) -> pd.DataFrame:
    keep = caliber.copy()
    keys = [
        "机构范围",
        "代码口径",
        "清理口径",
        "缺数案例处理",
        "低价值剔除",
        "历史百分位口径",
        "当前PE口径",
        "当前PB口径",
        "本次刷新范围",
    ]
    keep = keep[keep["口径项"].isin(keys)].copy()
    order = {k: i for i, k in enumerate(keys)}
    keep["_order"] = keep["口径项"].map(order)
    keep = keep.sort_values("_order").drop(columns=["_order"])
    return keep


def style_sheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    for col_name, col_idx in header_map.items():
        if col_name in PCT_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = "0.0%"
        elif col_name in MONEY_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = '#,##0.00_);(#,##0.00)'
        elif col_name in RATIO_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = '0.00_);(0.00)'
        elif col_name in DATE_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = "yyyy-mm-dd"

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).font = BOLD_FONT
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(vertical="center")

    for i, column in enumerate(df.columns, start=1):
        values = [str(column)]
        values.extend("" if v is None else str(v) for v in df[column].head(200))
        max_len = min(max(len(v) for v in values) + 2, 28)
        if column in {"当前持有机构", "机构口径", "估值口径备注", "备注"}:
            max_len = min(max_len + 8, 42)
        ws.column_dimensions[get_column_letter(i)].width = max_len

    if "是否满足主池完整度" in header_map:
        col = header_map["是否满足主池完整度"]
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, col).value
            if val is True or str(val).strip().lower() == "true":
                ws.cell(row, col).fill = PatternFill("solid", fgColor="E2F0D9")
            else:
                ws.cell(row, col).fill = PatternFill("solid", fgColor="FCE4D6")


def style_caliber_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).fill = SUB_FILL
        ws.cell(row, 1).font = BOLD_FONT
        ws.cell(row, 1).alignment = Alignment(vertical="top")
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 96


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame, caliber: bool = False) -> None:
    ws = wb.create_sheet(name)
    ws.append(df.columns.tolist())
    for row in df.replace({pd.NA: None}).itertuples(index=False, name=None):
        ws.append(list(row))
    if caliber:
        style_caliber_sheet(ws)
    else:
        style_sheet(ws, df)


def main() -> None:
    hot, val, caliber, decisions, shortlist, keep_missing, limited = load_frames()
    main_df = build_main_sheet(val, decisions)
    pref_df = build_preference_sheet(hot)
    val_df = build_valuation_aux(val)
    short_df = build_shortlist_sheet(shortlist, decisions)
    neg_df = build_negative_samples_sheet(keep_missing, limited)
    caliber_df = build_caliber_sheet(caliber)

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "研究主表", main_df)
    write_sheet(wb, "阶段短名单", short_df)
    write_sheet(wb, "观察与负样本", neg_df)
    write_sheet(wb, "历史偏好辅助", pref_df)
    write_sheet(wb, "估值辅助", val_df)
    write_sheet(wb, "口径说明", caliber_df, caliber=True)
    wb.save(OUTPUT)
    print(f"saved: {OUTPUT}")


if __name__ == "__main__":
    main()
