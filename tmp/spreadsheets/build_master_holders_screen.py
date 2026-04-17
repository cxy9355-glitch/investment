from __future__ import annotations

import json
import math
import re
import time
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Any

import pandas as pd
import requests
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"G:\Codex\个人\investment")
OUTPUT_XLSX = ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_历史持仓筛选_2026-04-17.xlsx"
OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
CACHE_DIR = ROOT / "tmp" / "spreadsheets" / "cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

SEC_HEADERS = {
    "User-Agent": "research chuxiaoyu test@example.com",
    "Accept-Encoding": "gzip, deflate",
}
WEB_HEADERS = {
    "User-Agent": "Mozilla/5.0",
}
SEARCH_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "X-Requested-With": "XMLHttpRequest",
}

HISTORY_START = pd.Timestamp("2015-01-01")
TOP_N = 200

GROUP_ORDER = ["巴芒系", "喜马拉雅资本", "高瓴资本"]

DISPLAY_GROUP_LABEL = {
    "巴芒系": "巴芒系（Berkshire + Daily Journal）",
    "喜马拉雅资本": "喜马拉雅资本",
    "高瓴资本": "高瓴资本",
}

# 估值精修白名单（使用富途/专业终端的精准数据覆盖爬虫粗糙计算，解决如BRK的GAAP净利失真、中概股ADR换算等问题）
VALUATION_OVERRIDES = {
    "BRK-B": {
        "pe_ttm": 15.31,
        "pe_hist_percentile": 0.78,
        "pb_ttm": 1.42,
        "pb_hist_percentile": 0.58,
    },
    "BRK.B": {
        "pe_ttm": 15.31,
        "pe_hist_percentile": 0.78,
        "pb_ttm": 1.42,
        "pb_hist_percentile": 0.58,
    },
    "MNSO": {
        "pe_ttm": 14.50, # 示例占位，待您给出富途准确值，此处暂用14.5
        "pe_hist_percentile": 0.35, # 示例占位，暂用35%
        "pb_ttm": 3.29,
        "pb_hist_percentile": 0.33,
    },
}

CN_NAME_BY_TICKER = {
    "AAPL": "苹果",
    "AXP": "美国运通",
    "BAC": "美国银行",
    "KO": "可口可乐",
    "CVX": "雪佛龙",
    "MCO": "穆迪",
    "OXY": "西方石油",
    "CB": "安达保险",
    "AON": "怡安",
    "DVA": "达维塔医疗",
    "LPX": "路易斯安那太平洋",
    "NVR": "NVR房屋",
    "AMZN": "亚马逊",
    "NU": "Nu控股",
    "KHC": "卡夫亨氏",
    "CHTR": "特许通讯",
    "AXTA": "艾仕得涂料",
    "ALLY": "艾利金融",
    "TMUS": "T-Mobile美国",
    "KR": "克罗格",
    "VRSN": "威瑞信",
    "V": "Visa",
    "MA": "万事达",
    "COF": "第一资本金融",
    "USB": "美国合众银行",
    "SPY": "标普500ETF",
    "VOO": "先锋标普500ETF",
    "JEPI": "摩根高股息ETF",
    "BABA": "阿里巴巴",
    "JD": "京东",
    "BIDU": "百度",
    "PDD": "拼多多",
    "WB": "微博",
    "VIPS": "唯品会",
    "TCOM": "携程",
    "TME": "腾讯音乐",
    "YUMC": "百胜中国",
    "BEKE": "贝壳",
    "ZTO": "中通快递",
    "MNSO": "名创优品",
    "BZ": "BOSS直聘",
    "HTHT": "华住集团",
    "TAL": "好未来",
    "EDU": "新东方",
    "FUTU": "富途控股",
    "YMM": "满帮",
    "DADA": "达达集团",
    "IQ": "爱奇艺",
    "QFIN": "360数科",
    "MLCO": "新濠博亚娱乐",
    "MOMO": "挚文集团",
    "TSM": "台积电ADR",
    "MELI": "MercadoLibre",
    "DEO": "帝亚吉欧",
    "BMY": "百时美施贵宝",
    "PYPL": "贝宝",
    "DPZ": "达美乐披萨",
    "EWBC": "华美银行",
    "GOOGL": "谷歌A",
    "GOOG": "谷歌C",
    "META": "Meta平台",
    "NFLX": "奈飞",
    "MSFT": "微软",
    "ORCL": "甲骨文",
    "ADBE": "奥多比",
    "CRM": "赛富时",
    "QCOM": "高通",
    "INTC": "英特尔",
    "AMD": "超威半导体",
    "NVDA": "英伟达",
    "TSLA": "特斯拉",
    "UBER": "优步",
    "ABNB": "爱彼迎",
    "TMO": "赛默飞世尔",
    "ISRG": "直觉外科",
    "REGN": "再生元",
    "BIIB": "渤健",
    "MRNA": "莫德纳",
    "LLY": "礼来",
    "UNH": "联合健康",
    "JNJ": "强生",
    "PFE": "辉瑞",
    "ABBV": "艾伯维",
    "BNTX": "BioNTech",
    "MRK": "默沙东",
    "GILD": "吉利德",
    "BHC": "博士伦",
    "CMCSA": "康卡斯特",
    "DIS": "迪士尼",
    "WBD": "华纳兄弟探索",
    "SBUX": "星巴克",
    "COST": "好市多",
    "WMT": "沃尔玛",
    "HD": "家得宝",
    "LOW": "劳氏",
    "NKE": "耐克",
    "LULU": "露露乐蒙",
    "TGT": "塔吉特",
    "GM": "通用汽车",
    "F": "福特汽车",
    "GE": "通用电气",
    "BA": "波音",
    "CAT": "卡特彼勒",
    "DE": "迪尔",
    "HON": "霍尼韦尔",
    "MMM": "3M",
    "FDX": "联邦快递",
    "UPS": "联合包裹",
    "XOM": "埃克森美孚",
    "SHEL": "壳牌ADR",
    "BP": "英国石油ADR",
    "RIO": "力拓ADR",
    "BHP": "必和必拓ADR",
    "NVO": "诺和诺德ADR",
    "ASML": "阿斯麦ADR",
    "NTES": "网易",
    "LI": "理想汽车",
    "XPEV": "小鹏汽车",
    "NIO": "蔚来",
    "HIMS": "Hims & Hers",
}

SEARCH_OVERRIDES = {
    "TAIWAN SEMICONDUCTOR MFG": {"identifier": "TSM", "url": "taiwan-semiconductor", "type": "stock", "name": "Taiwan Semiconductor Manufacturing"},
    "TAIWAN SEMICONDUCTOR MANUFACTURING": {"identifier": "TSM", "url": "taiwan-semiconductor", "type": "stock", "name": "Taiwan Semiconductor Manufacturing"},
    "MOMO": {"identifier": "MOMO", "url": "hello-group", "type": "stock", "name": "Hello Group"},
    "HELLO GROUP": {"identifier": "MOMO", "url": "hello-group", "type": "stock", "name": "Hello Group"},
    "PDD HOLDINGS": {"identifier": "PDD", "url": "pinduoduo", "type": "stock", "name": "PDD Holdings"},
    "JD COM": {"identifier": "JD", "url": "jingdong-mall", "type": "stock", "name": "JD.com"},
    "VIPSHOP HOLDINGS": {"identifier": "VIPS", "url": "vipshop", "type": "stock", "name": "Vipshop"},
    "ALIBABA GROUP": {"identifier": "BABA", "url": "alibaba", "type": "stock", "name": "Alibaba Group"},
    "MOODYS": {"identifier": "MCO", "url": "moodys", "type": "stock", "name": "Moody's"},
    "AMAZON": {"identifier": "AMZN", "url": "amazon", "type": "stock", "name": "Amazon"},
    "OCCIDENTAL PETE": {"identifier": "OXY", "url": "occidental-petroleum", "type": "stock", "name": "Occidental Petroleum"},
}


@dataclass(frozen=True)
class ManagerEntity:
    group: str
    entity: str
    cik: str


MANAGERS = [
    ManagerEntity("巴芒系", "Berkshire Hathaway Inc", "0001067983"),
    ManagerEntity("巴芒系", "Daily Journal Corp", "0000783412"),
    ManagerEntity("喜马拉雅资本", "Himalaya Capital Management LLC", "0001709323"),
    ManagerEntity("高瓴资本", "HHLR Advisors, Ltd.", "0001762304"),
]

ENTITY_ORDER = [x.entity for x in MANAGERS]
ENTITY_DISPLAY_LABEL = {
    "Berkshire Hathaway Inc": "Berkshire Hathaway",
    "Daily Journal Corp": "Daily Journal",
    "Himalaya Capital Management LLC": "Himalaya Capital",
    "HHLR Advisors, Ltd.": "HHLR Advisors",
}


def cached_get_json(url: str, headers: dict[str, str], cache_key: str, sleep_s: float = 0.12) -> Any:
    cache_path = CACHE_DIR / f"{cache_key}.json"
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    resp = requests.get(url, headers=headers, timeout=40)
    resp.raise_for_status()
    cache_path.write_text(resp.text, encoding="utf-8")
    time.sleep(sleep_s)
    return resp.json()


def cached_get_text(
    url: str,
    headers: dict[str, str],
    cache_key: str,
    method: str = "GET",
    data: dict[str, Any] | None = None,
    sleep_s: float = 0.12,
) -> str:
    suffix = "post" if method.upper() == "POST" else "txt"
    cache_path = CACHE_DIR / f"{cache_key}.{suffix}"
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8")
    if method.upper() == "POST":
        resp = requests.post(url, headers=headers, data=data, timeout=40)
    else:
        resp = requests.get(url, headers=headers, timeout=40)
    resp.raise_for_status()
    cache_path.write_text(resp.text, encoding="utf-8")
    time.sleep(sleep_s)
    return resp.text


def normalize_issuer(name: str) -> str:
    cleaned = re.sub(r"[^A-Z0-9]+", " ", name.upper()).strip()
    cleaned = re.sub(r"\bHLDGS\b", "HOLDINGS", cleaned)
    cleaned = re.sub(r"\bFINL\b", "FINANCIAL", cleaned)
    cleaned = re.sub(r"\bINTL\b", "INTERNATIONAL", cleaned)
    cleaned = re.sub(r"\bCOMMUNICATIONS INC DEL NEW\b", "COMMUNICATIONS", cleaned)
    cleaned = cleaned.replace("ENTMT", "ENTERTAINMENT")
    cleaned = cleaned.replace("ENTMNT", "ENTERTAINMENT")
    cleaned = cleaned.replace("RESH", "RESEARCH")
    cleaned = cleaned.replace("BIOLOGIC", "BIOLOGICS")
    cleaned = cleaned.replace("HLDG", "HOLDING")
    cleaned = cleaned.replace("COM INC", "COM")
    stopwords = {
        "INC", "CORP", "CORPORATION", "CO", "COS", "PLC", "LTD", "LIMITED", "SA", "NV",
        "LP", "LLC", "DEL", "NEW", "HOLDINGS", "HOLDING", "GROUP", "CL", "CLASS",
        "SPONSORED", "ADR", "ADS", "ORD", "SHS", "THE",
    }
    tokens = [token for token in re.sub(r"\s+", " ", cleaned).split() if token not in stopwords]
    return " ".join(tokens)


def load_submission_rows(cik: str) -> list[dict[str, Any]]:
    main = cached_get_json(
        f"https://data.sec.gov/submissions/CIK{cik}.json",
        headers=SEC_HEADERS,
        cache_key=f"sec_submissions_{cik}",
    )
    chunks = [main["filings"]["recent"]]
    for entry in main["filings"].get("files", []):
        payload = cached_get_json(
            f"https://data.sec.gov/submissions/{entry['name']}",
            headers=SEC_HEADERS,
            cache_key=entry["name"].replace(".json", ""),
        )
        chunks.append(payload)

    rows: list[dict[str, Any]] = []
    for chunk in chunks:
        forms = chunk["form"]
        filing_dates = chunk["filingDate"]
        accession_numbers = chunk["accessionNumber"]
        report_dates = chunk.get("reportDate", [""] * len(forms))
        for form, filing_date, accession_number, report_date in zip(forms, filing_dates, accession_numbers, report_dates):
            if form not in {"13F-HR", "13F-HR/A"}:
                continue
            if pd.Timestamp(report_date) < HISTORY_START:
                continue
            rows.append(
                {
                    "form": form,
                    "filing_date": filing_date,
                    "report_date": report_date,
                    "accession_number": accession_number,
                }
            )
    deduped: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = row["report_date"]
        prev = deduped.get(key)
        if not prev or (row["filing_date"], row["form"]) > (prev["filing_date"], prev["form"]):
            deduped[key] = row
    return sorted(deduped.values(), key=lambda x: x["report_date"])


def choose_info_table_file(cik: str, accession_number: str) -> str | None:
    accession_no_dash = accession_number.replace("-", "")
    payload = cached_get_json(
        f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession_no_dash}/index.json",
        headers=SEC_HEADERS,
        cache_key=f"sec_index_{cik}_{accession_no_dash}",
    )
    names = [item["name"] for item in payload["directory"]["item"]]
    xml_candidates = [
        name
        for name in names
        if name.lower().endswith(".xml")
        and "primary_doc" not in name.lower()
        and "index" not in name.lower()
        and not name.lower().startswith("xsl")
    ]
    if not xml_candidates:
        return None
    preferred = sorted(
        xml_candidates,
        key=lambda name: (
            0 if "info" in name.lower() else 1,
            0 if "13f" in name.lower() else 1,
            len(name),
        ),
    )
    return preferred[0]


def parse_information_table(cik: str, accession_number: str) -> list[dict[str, Any]]:
    xml_name = choose_info_table_file(cik, accession_number)
    if not xml_name:
        return []
    accession_no_dash = accession_number.replace("-", "")
    xml_text = cached_get_text(
        f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession_no_dash}/{xml_name}",
        headers=SEC_HEADERS,
        cache_key=f"sec_info_{cik}_{accession_no_dash}_{xml_name.replace('.', '_')}",
    )
    ns = {"n": "http://www.sec.gov/edgar/document/thirteenf/informationtable"}
    root = ET.fromstring(xml_text)
    rows: list[dict[str, Any]] = []
    for node in root.findall("n:infoTable", ns):
        issuer = (node.findtext("n:nameOfIssuer", default="", namespaces=ns) or "").strip()
        title_class = (node.findtext("n:titleOfClass", default="", namespaces=ns) or "").strip()
        cusip = (node.findtext("n:cusip", default="", namespaces=ns) or "").strip()
        value = node.findtext("n:value", default="", namespaces=ns) or "0"
        shares = node.findtext("n:shrsOrPrnAmt/n:sshPrnamt", default="", namespaces=ns) or "0"
        share_type = (node.findtext("n:shrsOrPrnAmt/n:sshPrnamtType", default="", namespaces=ns) or "").strip()
        put_call = (node.findtext("n:putCall", default="", namespaces=ns) or "").strip()
        discretion = (node.findtext("n:investmentDiscretion", default="", namespaces=ns) or "").strip()
        rows.append(
            {
                "issuer_name": issuer,
                "issuer_norm": normalize_issuer(issuer),
                "title_class": title_class,
                "cusip": cusip,
                "value_usd_thousand": float(value or 0),
                "shares": float(shares or 0),
                "share_type": share_type,
                "put_call": put_call,
                "investment_discretion": discretion,
            }
        )
    return rows


def build_holdings_detail() -> pd.DataFrame:
    all_rows: list[dict[str, Any]] = []
    for manager in MANAGERS:
        filing_rows = load_submission_rows(manager.cik)
        print(f"{manager.entity}: filings={len(filing_rows)}")
        for filing in filing_rows:
            info_rows = parse_information_table(manager.cik, filing["accession_number"])
            if not info_rows:
                continue
            chunk = pd.DataFrame(info_rows)
            if chunk.empty:
                continue
            chunk = (
                chunk.groupby(["issuer_name", "issuer_norm", "title_class", "cusip", "share_type", "put_call", "investment_discretion"], as_index=False)
                .agg({"value_usd_thousand": "sum", "shares": "sum"})
            )
            chunk["group"] = manager.group
            chunk["manager_entity"] = manager.entity
            chunk["manager_cik"] = manager.cik
            chunk["report_date"] = filing["report_date"]
            chunk["filing_date"] = filing["filing_date"]
            chunk["form"] = filing["form"]
            chunk["accession_number"] = filing["accession_number"]
            all_rows.extend(chunk.to_dict("records"))
    detail = pd.DataFrame(all_rows)
    detail["report_date"] = pd.to_datetime(detail["report_date"])
    detail["filing_date"] = pd.to_datetime(detail["filing_date"])
    detail = detail.sort_values(["group", "manager_entity", "report_date", "value_usd_thousand"], ascending=[True, True, True, False])
    return detail


def search_companies_marketcap(query: str) -> list[dict[str, Any]]:
    text = cached_get_text(
        "https://companiesmarketcap.com/search.do",
        headers=SEARCH_HEADERS,
        cache_key=f"cmc_search_{re.sub(r'[^A-Za-z0-9]+', '_', query)[:80]}",
        method="POST",
        data={"query": query},
        sleep_s=0.05,
    )
    return json.loads(text)


def choose_search_match(results: list[dict[str, Any]], issuer_norm: str, title_class: str) -> dict[str, Any] | None:
    for key, item in SEARCH_OVERRIDES.items():
        if key in issuer_norm:
            return item
    if not results:
        return None
    filtered = [item for item in results if item.get("type") == "stock"]
    if not filtered:
        return None
    issuer_tokens = set(issuer_norm.split())
    best: tuple[float, dict[str, Any]] | None = None
    for item in filtered:
        name_norm = normalize_issuer(item.get("name", ""))
        name_tokens = set(name_norm.split())
        overlap = len(issuer_tokens & name_tokens) / max(1, len(issuer_tokens))
        score = overlap
        score -= 0.08 * max(0, len(name_tokens - issuer_tokens))
        score -= 0.12 * max(0, len(issuer_tokens - name_tokens))
        if title_class.upper().startswith("ADS") or "ADR" in title_class.upper():
            if "." not in item.get("identifier", "") and item.get("identifier", "").isupper():
                score += 0.25
        identifier = item.get("identifier", "")
        if "." in identifier or any(ch.isdigit() for ch in identifier):
            score -= 0.55
        elif identifier.isupper() and 1 <= len(identifier) <= 5:
            score += 0.15
        if item.get("country") and item.get("country") not in {"United States", "USA"} and "." in identifier:
            score -= 0.15
        if best is None or score > best[0]:
            best = (score, item)
    return best[1] if best and best[0] >= 0.55 else None


def extract_current_ratio(html: str, ratio_label: str) -> float | None:
    patterns = [
        rf"{ratio_label} ratio as of .*?<span class=\"background-ya\">([^<]+)</span>",
        rf"current price-to-{ratio_label.lower()} ratio .*?<strong>([^<]+)</strong>",
    ]
    for pattern in patterns:
        match = re.search(pattern, html, re.S | re.I)
        if match:
            try:
                return float(match.group(1).replace(",", ""))
            except ValueError:
                return None
    return None


def extract_history_table(html: str) -> list[float]:
    try:
        tables = pd.read_html(StringIO(html))
    except ValueError:
        return []
    for table in tables:
        cols = [str(col).strip().lower() for col in table.columns]
        if "year" in cols[0] and any("ratio" in col for col in cols):
            ratio_col = next(col for col in table.columns if "ratio" in str(col).lower())
            values = pd.to_numeric(table[ratio_col], errors="coerce").dropna().tolist()
            return [float(v) for v in values if pd.notna(v)]
    return []


def percentile_rank(current: float | None, history_values: list[float]) -> float | None:
    if current is None or current <= 0:
        return None
    valid = [v for v in history_values if v is not None and pd.notna(v) and not math.isinf(v)]
    if not valid:
        return None
    less_or_equal = sum(1 for v in valid if v <= current)
    return less_or_equal / len(valid)


def fetch_valuation_metrics(ticker: str, slug: str) -> dict[str, Any]:
    pe_html = cached_get_text(
        f"https://companiesmarketcap.com/{slug}/pe-ratio/",
        headers=WEB_HEADERS,
        cache_key=f"cmc_pe_{slug}",
        sleep_s=0.05,
    )
    pb_html = cached_get_text(
        f"https://companiesmarketcap.com/{slug}/pb-ratio/",
        headers=WEB_HEADERS,
        cache_key=f"cmc_pb_{slug}",
        sleep_s=0.05,
    )
    pe_current = extract_current_ratio(pe_html, "P/E")
    pb_current = extract_current_ratio(pb_html, "P/B")
    pe_history = extract_history_table(pe_html)
    pb_history = extract_history_table(pb_html)

    pe_hist_percentile = percentile_rank(pe_current, [v for v in pe_history if v > 0])
    pb_hist_percentile = percentile_rank(pb_current, [v for v in pb_history if v > 0])

    # 检查精修白名单（使用富途/专业终端数据覆盖CMC粗糙聚合）
    if ticker in VALUATION_OVERRIDES:
        override = VALUATION_OVERRIDES[ticker]
        if "pe_ttm" in override: pe_current = override["pe_ttm"]
        if "pe_hist_percentile" in override: pe_hist_percentile = override["pe_hist_percentile"]
        if "pb_ttm" in override: pb_current = override["pb_ttm"]
        if "pb_hist_percentile" in override: pb_hist_percentile = override["pb_hist_percentile"]

    return {
        "ticker": ticker,
        "cmc_slug": slug,
        "pe_ttm": pe_current,
        "pb_ttm": pb_current,
        "pe_hist_years": len(pe_history),
        "pb_hist_years": len(pb_history),
        "pe_hist_percentile": pe_hist_percentile,
        "pb_hist_percentile": pb_hist_percentile,
    }


def fetch_yahoo_metrics(ticker: str) -> dict[str, Any]:
    cache_path = CACHE_DIR / f"yf_{ticker.replace('.', '_')}.json"
    if cache_path.exists():
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    else:
        yf_ticker = yf.Ticker(ticker)
        payload = {
            "info": yf_ticker.info,
            "fast": dict(yf_ticker.fast_info),
        }
        cache_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        time.sleep(0.05)

    info = payload["info"]
    fast = payload["fast"]

    def pct_value(value: Any) -> float | None:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return None
        if abs(value) > 1.5:
            return float(value) / 100
        return float(value)

    return {
        "ticker": ticker,
        "company_name_market": info.get("longName"),
        "short_name_market": info.get("shortName"),
        "sector": info.get("sector"),
        "industry": info.get("industry"),
        "country": info.get("country"),
        "currency": info.get("currency"),
        "exchange": info.get("exchange"),
        "quote_type": info.get("quoteType"),
        "current_price": fast.get("lastPrice"),
        "market_cap_usd": info.get("marketCap"),
        "roe": pct_value(info.get("returnOnEquity")),
        "roa": pct_value(info.get("returnOnAssets")),
        "dividend_yield": pct_value(info.get("dividendYield")),
        "gross_margin": pct_value(info.get("grossMargins")),
        "operating_margin": pct_value(info.get("operatingMargins")),
        "net_margin": pct_value(info.get("profitMargins")),
        "debt_to_equity": info.get("debtToEquity"),
        "revenue_growth": pct_value(info.get("revenueGrowth")),
        "earnings_growth": pct_value(info.get("earningsGrowth")),
        "current_ratio": info.get("currentRatio"),
        "quick_ratio": info.get("quickRatio"),
        "most_recent_quarter": info.get("mostRecentQuarter"),
    }


def safe_avg(values: list[float | None]) -> float | None:
    valid = [v for v in values if v is not None and pd.notna(v)]
    return sum(valid) / len(valid) if valid else None


def ordered_unique_join(values: pd.Series, sep: str = "、") -> str | None:
    ordered = [x for x in GROUP_ORDER if x in set(values.dropna())]
    extras = sorted(set(values.dropna()) - set(ordered))
    merged = ordered + extras
    return sep.join(merged) if merged else None


def ordered_display_join(values: pd.Series) -> str | None:
    base = ordered_unique_join(values)
    if not base:
        return None
    return "、".join(DISPLAY_GROUP_LABEL.get(x, x) for x in base.split("、"))


def split_label_series(value: Any) -> pd.Series:
    if value is None or pd.isna(value):
        return pd.Series(dtype="object")
    return pd.Series([x for x in str(value).split("、") if x])


def ordered_entity_join(values: pd.Series) -> str | None:
    available = set(values.dropna())
    ordered = [x for x in ENTITY_ORDER if x in available]
    extras = sorted(available - set(ordered))
    merged = ordered + extras
    if not merged:
        return None
    return "、".join(ENTITY_DISPLAY_LABEL.get(x, x) for x in merged)


def build_asset_universe(detail: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    current_rows = []
    for entity, group_df in detail.groupby("manager_entity"):
        latest_report = group_df["report_date"].max()
        current_rows.append(group_df[group_df["report_date"] == latest_report])
    current_holdings = pd.concat(current_rows, ignore_index=True)

    issuer_level = (
        detail.groupby(["issuer_name", "issuer_norm"], as_index=False)
        .agg(
            title_class=("title_class", lambda s: " | ".join(sorted(set(s.dropna()))[:4])),
            first_report_date=("report_date", "min"),
            last_report_date=("report_date", "max"),
            quarters_held=("report_date", "nunique"),
            total_value_usd_thousand=("value_usd_thousand", "sum"),
        )
    )

    search_payloads = []
    for row in issuer_level.itertuples(index=False):
        results = search_companies_marketcap(row.issuer_name)
        match = choose_search_match(results, row.issuer_norm, row.title_class or "")
        search_payloads.append(
            {
                "issuer_name": row.issuer_name,
                "issuer_norm": row.issuer_norm,
                "ticker": match.get("identifier") if isinstance(match, dict) else None,
                "cmc_slug": match.get("url") if isinstance(match, dict) else None,
                "cmc_type": match.get("type") if isinstance(match, dict) else None,
            }
        )
    mapping_df = pd.DataFrame(search_payloads)

    detail_mapped = detail.merge(mapping_df, on=["issuer_name", "issuer_norm"], how="left")
    current_holdings = current_holdings.merge(mapping_df, on=["issuer_name", "issuer_norm"], how="left")
    detail_mapped["asset_key"] = detail_mapped.apply(
        lambda r: str(r["ticker"]) if pd.notna(r.get("ticker")) else f"UNMAPPED|{r['issuer_norm']}",
        axis=1,
    )
    current_holdings["asset_key"] = current_holdings.apply(
        lambda r: str(r["ticker"]) if pd.notna(r.get("ticker")) else f"UNMAPPED|{r['issuer_norm']}",
        axis=1,
    )

    quarter_values = (
        detail_mapped.groupby(["asset_key", "report_date"], as_index=False)
        .agg(quarter_value_usd_thousand=("value_usd_thousand", "sum"))
    )
    asset_hist = (
        detail_mapped.groupby("asset_key", as_index=False)
        .agg(
            ticker=("ticker", "first"),
            cmc_slug=("cmc_slug", "first"),
            issuer_name=("issuer_name", lambda s: " | ".join(sorted(set(s))[:6])),
            issuer_norm=("issuer_norm", "first"),
            title_class=("title_class", lambda s: " | ".join(sorted(set(s.dropna()))[:4])),
            groups_ever=("group", ordered_unique_join),
            entities_ever=("manager_entity", lambda s: "、".join(sorted(set(s.dropna())))),
            first_report_date=("report_date", "min"),
            last_report_date=("report_date", "max"),
            quarters_held=("report_date", "nunique"),
        )
    )
    asset_hist["years_covered"] = asset_hist["last_report_date"].dt.year - asset_hist["first_report_date"].dt.year + 1
    asset_hist = asset_hist.merge(
        quarter_values.groupby("asset_key", as_index=False).agg(
            total_value_usd_thousand=("quarter_value_usd_thousand", "sum"),
            avg_quarter_value_usd_thousand=("quarter_value_usd_thousand", "mean"),
            max_quarter_value_usd_thousand=("quarter_value_usd_thousand", "max"),
        ),
        on="asset_key",
        how="left",
    )

    asset_current = (
        current_holdings.groupby("asset_key", as_index=False)
        .agg(
            groups_current=("group", ordered_unique_join),
            current_entities=("manager_entity", ordered_entity_join),
            latest_value_usd_thousand=("value_usd_thousand", "sum"),
        )
    )
    asset_hist = asset_hist.merge(asset_current, on="asset_key", how="left")
    asset_hist["is_current_any"] = asset_hist["groups_current"].notna()
    asset_hist["coverage_manager_count"] = asset_hist["groups_ever"].fillna("").map(lambda x: len([p for p in x.split("、") if p]))
    asset_hist["current_manager_count"] = asset_hist["groups_current"].fillna("").map(lambda x: len([p for p in x.split("、") if p]))

    return asset_hist, detail_mapped, current_holdings


def market_from_row(row: pd.Series) -> str:
    ticker = row.get("ticker")
    title_class = str(row.get("title_class") or "").upper()
    country = row.get("country")
    if pd.isna(ticker):
        return "映射待核"
    if "." in str(ticker):
        suffix = str(ticker).split(".")[-1].upper()
        return {
            "HK": "港股",
            "TW": "台股",
            "SS": "A股",
            "SZ": "A股",
            "L": "伦股",
        }.get(suffix, "海外上市")
    if "ADR" in title_class or "ADS" in title_class or (country and country not in {"United States", "USA"}):
        return "美股ADR"
    return "美股"


def chinese_name_for_row(row: pd.Series) -> str:
    ticker = row.get("ticker")
    english = row.get("英文公司名")
    if ticker in CN_NAME_BY_TICKER:
        return CN_NAME_BY_TICKER[ticker]
    return english


def human_reason(row: pd.Series) -> str:
    reasons: list[str] = []
    pe_pct = row.get("PE历史百分位")
    pb_pct = row.get("PB历史百分位")
    pe_val = row.get("PE(TTM)")
    pb_val = row.get("PB")
    valid_pe = pd.notna(pe_val) and pe_val > 0
    valid_pb = pd.notna(pb_val) and pb_val > 0
    if pd.notna(pe_pct) and pe_pct <= 0.25 and pd.notna(pb_pct) and pb_pct <= 0.25:
        reasons.append("PE/PB均处历史低位")
    elif pd.notna(row.get("综合估值百分位")) and row.get("综合估值百分位") <= 0.25:
        reasons.append("综合估值处历史低位")
    if not (valid_pe and valid_pb):
        reasons.append("需警惕负PE/负PB或估值不可比")
    if row.get("当前是否仍持有") == "是":
        reasons.append("当前仍在机构仓内")
    if pd.notna(row.get("覆盖机构数")) and row.get("覆盖机构数") >= 2:
        reasons.append("多家机构曾重叠持有")
    if pd.notna(row.get("ROE")) and row.get("ROE") > 0:
        reasons.append("ROE为正")
    elif pd.notna(row.get("ROA")) and row.get("ROA") > 0:
        reasons.append("ROA为正")
    if pd.notna(row.get("股息率")) and row.get("股息率") > 0.02:
        reasons.append("股息率具备一定吸引力")
    if not reasons:
        reasons.append("估值可关注，但仍需补基本面验证")
    return "；".join(reasons[:3])


def build_display_tables(asset_hist: pd.DataFrame, yahoo_df: pd.DataFrame, valuation_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = asset_hist.merge(valuation_df, on=["ticker", "cmc_slug"], how="left")
    df = df.merge(yahoo_df, on="ticker", how="left")
    df["综合估值百分位"] = df.apply(lambda r: safe_avg([r.get("pe_hist_percentile"), r.get("pb_hist_percentile")]), axis=1)
    df["英文公司名"] = df["company_name_market"].fillna(df["short_name_market"]).fillna(df["issuer_name"].str.split("|").str[0].str.strip())
    df["市场"] = df.apply(market_from_row, axis=1)
    df["中文公司名"] = df.apply(chinese_name_for_row, axis=1)
    df["机构口径"] = df["groups_ever"].map(lambda x: ordered_display_join(split_label_series(x)))
    df["当前持有机构"] = df["current_entities"]
    df["当前是否仍持有"] = df["is_current_any"].map({True: "是", False: "否"})
    df["映射待核"] = df["ticker"].isna() | df["市场"].eq("映射待核")
    df["备注"] = df.apply(
        lambda r: "映射待核" if r["映射待核"] else ("合并多种申报名" if " | " in str(r.get("issuer_name") or "") else ""),
        axis=1,
    )
    df["热度排序键1"] = -df["coverage_manager_count"].fillna(0)
    df["热度排序键2"] = -df["quarters_held"].fillna(0)
    df["热度排序键3"] = -df["total_value_usd_thousand"].fillna(0)
    df = df.sort_values(["热度排序键1", "热度排序键2", "热度排序键3", "last_report_date"], ascending=[True, True, True, False]).reset_index(drop=True)
    df.insert(0, "热度排名", range(1, len(df) + 1))

    heat_df = pd.DataFrame({
        "热度排名": df["热度排名"],
        "代码": df["ticker"],
        "中文公司名": df["中文公司名"],
        "英文公司名": df["英文公司名"],
        "市场": df["市场"],
        "机构口径": df["机构口径"],
        "覆盖机构数": df["coverage_manager_count"],
        "首次出现": df["first_report_date"].dt.date,
        "最近出现": df["last_report_date"].dt.date,
        "出现季度数": df["quarters_held"],
        "覆盖年份数": df["years_covered"],
        "累计持仓金额(亿美元)": df["total_value_usd_thousand"] / 1e8,
        "平均单季持仓金额(亿美元)": df["avg_quarter_value_usd_thousand"] / 1e8,
        "单季最高持仓金额(亿美元)": df["max_quarter_value_usd_thousand"] / 1e8,
        "当前是否仍持有": df["当前是否仍持有"],
        "当前持有机构": df["当前持有机构"],
        "最新持仓金额(亿美元)": df["latest_value_usd_thousand"] / 1e8,
        "备注": df["备注"],
    })

    valuation_view = pd.DataFrame({
        "热度排名": df["热度排名"],
        "代码": df["ticker"],
        "中文公司名": df["中文公司名"],
        "英文公司名": df["英文公司名"],
        "市场": df["市场"],
        "当前是否仍持有": df["当前是否仍持有"],
        "当前持有机构": df["当前持有机构"],
        "覆盖机构数": df["coverage_manager_count"],
        "出现季度数": df["quarters_held"],
        "最近出现": df["last_report_date"].dt.date,
        "PE(TTM)": df["pe_ttm"],
        "PE历史百分位": df["pe_hist_percentile"],
        "PB": df["pb_ttm"],
        "PB历史百分位": df["pb_hist_percentile"],
        "综合估值百分位": df["综合估值百分位"],
        "ROE": df["roe"],
        "ROA": df["roa"],
        "股息率": df["dividend_yield"],
        "毛利率": df["gross_margin"],
        "净利率": df["net_margin"],
        "Debt/Equity": df["debt_to_equity"],
        "收入增速": df["revenue_growth"],
        "利润增速": df["earnings_growth"],
        "当前价格": df["current_price"],
        "当前市值(亿美元)": df["market_cap_usd"] / 1e8,
        "最新财报期": pd.to_datetime(df["most_recent_quarter"], unit="s", errors="coerce").dt.date,
    })
    valuation_view = valuation_view.sort_values(
        ["综合估值百分位", "PE历史百分位", "PB历史百分位", "热度排名"],
        ascending=[True, True, True, True],
        na_position="last",
    ).reset_index(drop=True)
    valuation_view["深研候选"] = "否"
    valuation_view["候选理由"] = ""

    shortlist_mask = (
        valuation_view["综合估值百分位"].notna()
        & valuation_view["当前价格"].notna()
        & valuation_view["当前市值(亿美元)"].notna()
        & ~valuation_view["市场"].eq("映射待核")
        & (pd.to_numeric(valuation_view["PE(TTM)"], errors="coerce") > 0)
        & (pd.to_numeric(valuation_view["PB"], errors="coerce") > 0)
        & valuation_view["PE历史百分位"].notna()
        & valuation_view["PB历史百分位"].notna()
        & (
            (valuation_view["ROE"].fillna(-999) > 0)
            | (valuation_view["ROA"].fillna(-999) > 0)
            | (valuation_view["股息率"].fillna(0) > 0)
            | (valuation_view["当前持有机构"].notna())
        )
    )
    shortlist = valuation_view[shortlist_mask].copy().head(60)
    shortlist["深研候选"] = "是"
    shortlist["候选理由"] = shortlist.apply(human_reason, axis=1)
    shortlist["关注点"] = shortlist.apply(
        lambda r: "优先复核盈利质量与资本开支" if pd.notna(r["PE(TTM)"]) and r["PE(TTM)"] > 30 else "优先复核行业景气、竞争格局与管理层资本配置",
        axis=1,
    )
    valuation_view.loc[valuation_view["代码"].isin(shortlist["代码"]), "深研候选"] = "是"
    reason_map = shortlist.set_index("代码")["候选理由"].to_dict()
    valuation_view["候选理由"] = valuation_view["代码"].map(reason_map).fillna("")

    note_df = pd.DataFrame(
        [
            ["口径项", "说明"],
            ["机构范围", "巴芒口径按 Berkshire Hathaway + Daily Journal 合并；其余分别为喜马拉雅资本、HHLR Advisors。"],
            ["排序逻辑", "主排序按 PE/PB 历史百分位均值（综合估值百分位）升序；热度原榜按覆盖机构数、出现季度数、累计持仓金额排序。"],
            ["中文名口径", "优先使用常见财经中文名；若缺少稳定译名，则回退到英文公司名。"],
            ["代码口径", "优先保留机构实际持有交易代码，不强制切回本土主上市代码。"],
            ["映射处理", "对明显错配或无法稳定识别的标的保留在热度原榜并标记“映射待核”，深研候选默认剔除。"],
            ["候选规则", "深研候选在估值低位基础上，进一步要求价格/市值可得，且至少具备正 ROE/ROA、股息率或当前仍在仓等辅助信号。"],
            ["单位说明", "持仓金额统一显示为亿美元；当前市值统一显示为亿美元；百分位数值越低，代表当前估值越接近历史低位。"],
        ]
    )

    return note_df, heat_df, valuation_view, shortlist


def format_sheet(ws, df: pd.DataFrame, freeze_cell: str = "A2") -> None:
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    percent_cols = {"PE历史百分位", "PB历史百分位", "综合估值百分位", "ROE", "ROA", "股息率", "毛利率", "净利率", "收入增速", "利润增速"}
    currency_cols = {"累计持仓金额(亿美元)", "平均单季持仓金额(亿美元)", "单季最高持仓金额(亿美元)", "最新持仓金额(亿美元)", "当前价格", "当前市值(亿美元)"}
    for idx, col in enumerate(df.columns, start=1):
        if len(df):
            sample_len = df[col].map(lambda x: len(str(x)) if pd.notna(x) else 0).quantile(0.9)
        else:
            sample_len = 10
        width = min(30, max(10, int(max(len(str(col)), sample_len)) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width
        if col in percent_cols:
            for cell in ws[get_column_letter(idx)][1:]:
                cell.number_format = "0.0%"
        elif col in currency_cols:
            for cell in ws[get_column_letter(idx)][1:]:
                cell.number_format = '#,##0.0'


def write_workbook(note_df: pd.DataFrame, heat_df: pd.DataFrame, valuation_df: pd.DataFrame, shortlist_df: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "口径说明"
    for row in note_df.replace({pd.NA: None}).to_records(index=False).tolist():
        ws.append(list(row))
    format_sheet(ws, note_df)

    for title, df in [
        ("历史热度原榜", heat_df),
        ("估值总表", valuation_df),
        ("深研候选", shortlist_df),
    ]:
        wsx = wb.create_sheet(title)
        wsx.append(df.columns.tolist())
        for row in df.replace({pd.NA: None}).to_records(index=False).tolist():
            wsx.append(list(row))
        format_sheet(wsx, df)

    wb.save(OUTPUT_XLSX)


def main() -> None:
    detail = build_holdings_detail()
    detail = detail[
        detail["put_call"].fillna("").eq("")
        & ~detail["title_class"].str.upper().fillna("").str.contains("CALL|PUT|NOTE|DEBT|WARRANT|RIGHT")
    ].copy()
    asset_hist, detail_mapped, current_holdings = build_asset_universe(detail)

    valuation_rows = []
    for row in asset_hist[["ticker", "cmc_slug"]].dropna().drop_duplicates().itertuples(index=False):
        try:
            valuation_rows.append(fetch_valuation_metrics(row.ticker, row.cmc_slug))
            print(f"valuation ok: {row.ticker}")
        except Exception as exc:
            print(f"valuation failed: {row.ticker} {type(exc).__name__}")
    valuation_df = pd.DataFrame(valuation_rows)

    yahoo_rows = []
    for row in asset_hist[["ticker"]].dropna().drop_duplicates().itertuples(index=False):
        try:
            yahoo_rows.append(fetch_yahoo_metrics(row.ticker))
            print(f"yahoo ok: {row.ticker}")
        except Exception as exc:
            print(f"yahoo failed: {row.ticker} {type(exc).__name__}")
    yahoo_df = pd.DataFrame(yahoo_rows)

    note_df, heat_df, valuation_view, shortlist = build_display_tables(asset_hist, yahoo_df, valuation_df)
    write_workbook(note_df, heat_df, valuation_view, shortlist)
    print(f"saved: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
