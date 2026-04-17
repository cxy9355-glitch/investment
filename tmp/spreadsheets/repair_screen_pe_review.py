from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"G:\Codex\个人\investment")
WORKBOOKS = [
    ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_历史持仓筛选_2026-04-17_可比清理版.xlsx",
    ROOT / "机构持仓研究" / "archive_old_logs" / "巴芒_喜马拉雅_高瓴_历史持仓筛选_2026-04-17.xlsx",
]

HEADER_RENAMES = {
    "PE历史百分位": "PE年末历史百分位(自算)",
    "PB历史百分位": "PB年末历史百分位(自算)",
    "综合估值百分位": "综合估值百分位(自算)",
}

CORRECTIONS = {
    "BRK-B": {
        "pe_ttm": 15.31,
        "pe_hist_pct": 0.391304347826087,
        "deep_flag": "否",
        "reason": "当前PE已按Yahoo/Futu一致口径修正；自算年表分位不属历史低位",
    },
    "MNSO": {
        "pe_ttm": 27.84,
        "pe_hist_pct": 0.6666666666666666,
        "deep_flag": "否",
        "reason": "ADR每ADS=4股；原Yahoo/CMC口径低估PE，已按复核口径修正",
    },
}

CALIBER_ROWS = [
    ("PE(TTM)复核口径", "普通美股优先参考Yahoo与富途一致值；已抽样复核BRK-B。"),
    ("ADR估值修正规则", "ADR若出现EPS基准与报价单位不一致，优先按复核后的ADS口径修正；已抽样复核MNSO。"),
    ("历史百分位口径", "PE/PB历史百分位为基于CompaniesMarketCap年末历史表的自算分位，不等同于富途滚动历史百分位。"),
]


def safe_avg(values: list[float | None]) -> float | None:
    valid = [v for v in values if v is not None]
    return sum(valid) / len(valid) if valid else None


def rename_headers(ws) -> dict[str, int]:
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        new_value = HEADER_RENAMES.get(value, value)
        ws.cell(1, col).value = new_value
        headers[new_value] = col
    return headers


def patch_caliber(ws) -> None:
    existing = {
        (ws.cell(row, 1).value, ws.cell(row, 2).value)
        for row in range(2, ws.max_row + 1)
    }
    for item, note in CALIBER_ROWS:
        if (item, note) not in existing:
            ws.append([item, note])


def patch_valuation_sheet(ws) -> None:
    headers = rename_headers(ws)
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row, headers["代码"]).value
        if code not in CORRECTIONS:
            continue
        rule = CORRECTIONS[code]
        ws.cell(row, headers["PE(TTM)"]).value = rule["pe_ttm"]
        ws.cell(row, headers["PE年末历史百分位(自算)"]).value = rule["pe_hist_pct"]
        pb_pct = ws.cell(row, headers["PB年末历史百分位(自算)"]).value
        ws.cell(row, headers["综合估值百分位(自算)"]).value = safe_avg([rule["pe_hist_pct"], pb_pct])
        if "深研候选" in headers:
            ws.cell(row, headers["深研候选"]).value = rule["deep_flag"]
        if "候选理由" in headers:
            ws.cell(row, headers["候选理由"]).value = rule["reason"]


def patch_shortlist_sheet(ws) -> None:
    headers = rename_headers(ws)
    code_col = headers["代码"]
    remove_rows = []
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row, code_col).value
        if code in CORRECTIONS and CORRECTIONS[code]["deep_flag"] != "是":
            remove_rows.append(row)
    for row in reversed(remove_rows):
        ws.delete_rows(row, 1)


def patch_workbook(path: Path) -> None:
    if not path.exists():
        return
    wb = load_workbook(path)
    if "口径说明" in wb.sheetnames:
        patch_caliber(wb["口径说明"])
    if "估值总表" in wb.sheetnames:
        patch_valuation_sheet(wb["估值总表"])
    if "深研候选" in wb.sheetnames:
        patch_shortlist_sheet(wb["深研候选"])
    wb.save(path)
    print(f"patched: {path}")


def main() -> None:
    for path in WORKBOOKS:
        patch_workbook(path)


if __name__ == "__main__":
    main()
