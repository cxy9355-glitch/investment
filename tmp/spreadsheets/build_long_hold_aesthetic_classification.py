from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"G:\Codex\个人\investment")
DATASET = ROOT / "机构持仓研究" / "长期持有全池数据集_2026-04-17_可比清理版.xlsx"
RESEARCH = ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_研究总表_2026-04-17.xlsx"
OUTPUT = ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_长期持有审美分类总表_2026-04-17.xlsx"

# Stable model: durable facts live in DATASET standard sheets. RESEARCH is a
# derived reading workbook kept only as an auxiliary input for current display.
HOLDING_SHEETS = ["holding_timeline", "实体级持仓历史_可比池"]
OPERATING_SHEETS = ["operating_timeline", "年度经营质量_可比池"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)

PCT_COLS = {
    "平均持仓权重",
    "峰值持仓权重",
    "5年平均ROE",
    "5年平均ROA",
    "5年平均毛利率",
    "5年平均净利率",
    "5年平均FCF利润率",
    "持有持续性分位",
    "经营持续性分位",
}
SCORE_COLS = {"持有持续性分数", "经营质量水平分", "经营质量稳定分", "经营持续性分数"}
DATE_COLS = {"最近出现", "最新财报期"}


def percentile_rank(series: pd.Series, inverse: bool = False) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if inverse:
        numeric = -numeric
    return numeric.rank(pct=True)


def read_first_sheet(path: Path, sheet_names: list[str]) -> pd.DataFrame:
    available = set(pd.ExcelFile(path).sheet_names)
    for sheet_name in sheet_names:
        if sheet_name in available:
            return pd.read_excel(path, sheet_name=sheet_name)
    raise ValueError(f"None of {sheet_names} found in {path}")


def load_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    main = pd.read_excel(RESEARCH, sheet_name="研究主表")
    negative = pd.read_excel(RESEARCH, sheet_name="观察与负样本")
    holdings = read_first_sheet(DATASET, HOLDING_SHEETS)
    annual = read_first_sheet(DATASET, OPERATING_SHEETS)
    return main, negative, holdings, annual


def compute_holding_scores(holdings: pd.DataFrame) -> pd.DataFrame:
    work = holdings.copy()
    work["report_date"] = pd.to_datetime(work["report_date"])
    grouped = work.groupby("ticker").agg(
        覆盖机构数=("manager_entity", "nunique"),
        出现季度数=("report_date", "nunique"),
        覆盖年份数=("report_date", lambda s: s.dt.year.nunique()),
        平均持仓权重=("position_weight", "mean"),
        峰值持仓权重=("position_weight", "max"),
        当前持有机构数=("is_current", lambda s: int(pd.Series(s).fillna(False).sum())),
        当前是否仍持有=("is_current", lambda s: bool(pd.Series(s).fillna(False).any())),
    )
    grouped["出现季度数_分位"] = percentile_rank(grouped["出现季度数"])
    grouped["覆盖年份数_分位"] = percentile_rank(grouped["覆盖年份数"])
    grouped["覆盖机构数_分位"] = percentile_rank(grouped["覆盖机构数"])
    grouped["平均持仓权重_分位"] = percentile_rank(grouped["平均持仓权重"])
    grouped["峰值持仓权重_分位"] = percentile_rank(grouped["峰值持仓权重"])
    grouped["当前是否仍持有_分位"] = grouped["当前是否仍持有"].astype(int)
    grouped["持有持续性分数"] = (
        grouped["出现季度数_分位"] * 30
        + grouped["覆盖年份数_分位"] * 20
        + grouped["覆盖机构数_分位"] * 20
        + grouped["平均持仓权重_分位"] * 15
        + grouped["峰值持仓权重_分位"] * 10
        + grouped["当前是否仍持有_分位"] * 5
    )
    grouped["持有持续性分位"] = percentile_rank(grouped["持有持续性分数"])
    return grouped.reset_index()


def _recent_window(group: pd.DataFrame, years: int = 5) -> pd.DataFrame:
    return group.sort_values("fiscal_year").tail(min(years, len(group))).copy()


def _mean_metric(frame: pd.DataFrame, col: str) -> float:
    series = pd.to_numeric(frame[col], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    return float(series.mean()) if len(series) else np.nan


def _std_metric(frame: pd.DataFrame, col: str) -> float:
    series = pd.to_numeric(frame[col], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    return float(series.std(ddof=0)) if len(series) >= 2 else np.nan


def compute_operating_scores(annual: pd.DataFrame) -> pd.DataFrame:
    work = annual[annual["核心字段完整"] == True].copy()
    rows: list[dict[str, object]] = []
    for ticker, group in work.groupby("ticker"):
        recent = _recent_window(group)
        row = {
            "ticker": ticker,
            "有效财年行数": int(group["fiscal_year"].nunique()),
            "5年平均ROE": _mean_metric(recent, "roe"),
            "5年平均ROA": _mean_metric(recent, "roa"),
            "5年平均毛利率": _mean_metric(recent, "gross_margin"),
            "5年平均净利率": _mean_metric(recent, "net_margin"),
            "5年平均FCF利润率": _mean_metric(recent, "fcf_margin"),
            "5年平均Debt/Equity": _mean_metric(recent, "debt_to_equity"),
            "ROE波动": _std_metric(recent, "roe"),
            "净利率波动": _std_metric(recent, "net_margin"),
            "毛利率波动": _std_metric(recent, "gross_margin"),
            "Debt/Equity波动": _std_metric(recent, "debt_to_equity"),
            "净利润为正年份占比": float(recent["net_income"].gt(0).mean()) if len(recent) else np.nan,
            "FCF为正年份占比": float(recent["fcf"].dropna().gt(0).mean()) if recent["fcf"].notna().any() else np.nan,
        }
        rows.append(row)

    scored = pd.DataFrame(rows)
    level_cols = ["5年平均ROE", "5年平均ROA", "5年平均毛利率", "5年平均净利率", "5年平均FCF利润率"]
    for col in level_cols:
        scored[f"{col}_分位"] = percentile_rank(scored[col])
    scored["5年平均Debt/Equity_分位"] = percentile_rank(scored["5年平均Debt/Equity"], inverse=True)

    stability_cols = ["ROE波动", "净利率波动", "毛利率波动", "Debt/Equity波动"]
    for col in stability_cols:
        scored[f"{col}_分位"] = percentile_rank(scored[col], inverse=True)
    scored["净利润为正年份占比_分位"] = percentile_rank(scored["净利润为正年份占比"])
    scored["FCF为正年份占比_分位"] = percentile_rank(scored["FCF为正年份占比"])

    level_rank_cols = [f"{col}_分位" for col in level_cols] + ["5年平均Debt/Equity_分位"]
    stability_rank_cols = [f"{col}_分位" for col in stability_cols] + ["净利润为正年份占比_分位", "FCF为正年份占比_分位"]
    scored["经营质量水平分"] = scored[level_rank_cols].mean(axis=1, skipna=True) * 100
    scored["经营质量稳定分"] = scored[stability_rank_cols].mean(axis=1, skipna=True) * 100
    scored["经营持续性分数"] = scored["经营质量水平分"] * 0.6 + scored["经营质量稳定分"] * 0.4
    scored["经营持续性分位"] = percentile_rank(scored["经营持续性分数"])
    return scored


def classify_samples(frame: pd.DataFrame) -> pd.DataFrame:
    work = frame.copy()
    reasons: list[str] = []
    categories: list[str] = []
    for _, row in work.iterrows():
        obs_reasons: list[str] = []
        valid_years = row.get("有效财年行数", np.nan)
        hold_pct = row.get("持有持续性分位", np.nan)
        op_pct = row.get("经营持续性分位", np.nan)
        if pd.isna(valid_years) or valid_years < 5:
            obs_reasons.append("有效财年不足5年")
        if pd.isna(op_pct):
            obs_reasons.append("经营持续性分数缺失")
        if pd.notna(hold_pct) and 0.35 <= hold_pct <= 0.65:
            obs_reasons.append("持有持续性位于观察区")
        if pd.notna(op_pct) and 0.35 <= op_pct <= 0.65:
            obs_reasons.append("经营持续性位于观察区")

        if obs_reasons:
            categories.append("观察区")
            reasons.append("；".join(obs_reasons))
            continue

        if hold_pct >= 0.65 and op_pct >= 0.65:
            categories.append("长期赢家")
            reasons.append("持有与经营持续性均处高位")
        elif hold_pct >= 0.65 and op_pct <= 0.35:
            categories.append("审美失效/退化案例")
            reasons.append("长期被偏爱，但经营持续性已落入低位")
        elif hold_pct <= 0.35 and op_pct >= 0.65:
            categories.append("经营优秀但未被长期拿住")
            reasons.append("经营持续性高，但持有持续性偏低")
        elif hold_pct <= 0.35 and op_pct <= 0.35:
            categories.append("弱质/陷阱样本")
            reasons.append("持有与经营持续性均偏弱")
        else:
            categories.append("观察区")
            reasons.append("分位落在边界区间")

    work["分类结果"] = categories
    work["观察原因"] = reasons
    return work


def build_output_frames() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    main, negative, holdings, annual = load_inputs()
    hold_scores = compute_holding_scores(holdings)
    op_scores = compute_operating_scores(annual)

    hold_keep = hold_scores[
        [
            "ticker",
            "覆盖年份数",
            "当前持有机构数",
            "平均持仓权重",
            "峰值持仓权重",
            "持有持续性分数",
            "持有持续性分位",
        ]
    ].copy()
    merged = main.merge(hold_keep, left_on="代码", right_on="ticker", how="left")
    merged = merged.merge(op_scores, left_on="代码", right_on="ticker", how="left", suffixes=("", "_经营"))
    merged = classify_samples(merged)
    merged["当前持有状态"] = merged["当前持有机构"].apply(lambda x: "当前在仓" if pd.notna(x) and str(x).strip() else "当前不在仓")
    cols = [
        "代码", "中文公司名", "英文公司名", "市场", "当前持有机构", "当前持有状态",
        "覆盖机构数", "出现季度数", "覆盖年份数", "当前持有机构数", "平均持仓权重", "峰值持仓权重",
        "最新财报期", "有效财年行数", "核心完整财年行数", "是否满足主池完整度",
        "5年平均ROE", "5年平均ROA", "5年平均毛利率", "5年平均净利率", "5年平均FCF利润率", "5年平均Debt/Equity",
        "持有持续性分数", "持有持续性分位", "经营质量水平分", "经营质量稳定分", "经营持续性分数", "经营持续性分位",
        "分类结果", "观察原因",
    ]
    total = merged[cols].copy().sort_values(["分类结果", "持有持续性分数", "经营持续性分数"], ascending=[True, False, False]).reset_index(drop=True)

    summary = total.groupby("分类结果").agg(
        样本数=("代码", "count"),
        平均持有持续性分数=("持有持续性分数", "mean"),
        平均经营持续性分数=("经营持续性分数", "mean"),
    ).reset_index().sort_values("样本数", ascending=False)

    method = pd.DataFrame(
        [
            ["持有持续性输入", "覆盖机构数、出现季度数、覆盖年份数、平均持仓权重、峰值持仓权重、当前是否仍持有。"],
            ["持有持续性计算式", "先对每个输入字段在全样本内做分位归一化，得到0到1之间的分位值；再按公式：持有持续性分数 = 出现季度数分位×30 + 覆盖年份数分位×20 + 覆盖机构数分位×20 + 平均持仓权重分位×15 + 峰值持仓权重分位×10 + 当前是否仍持有×5。"],
            ["经营持续性输入", "基于近5个有效财年，计算5年平均ROE/ROA/毛利率/净利率/FCF利润率/Debt/Equity，以及波动与正收益年份占比。"],
            ["经营质量水平分计算式", "先对5年平均ROE、ROA、毛利率、净利率、FCF利润率做全样本分位归一化，对5年平均Debt/Equity做反向分位归一化；再取这6项分位的平均值并乘100。"],
            ["经营质量稳定分计算式", "先对ROE波动、净利率波动、毛利率波动、Debt/Equity波动做反向分位归一化，再加上净利润为正年份占比、FCF为正年份占比的正向分位；取这6项分位平均值并乘100。"],
            ["经营持续性合成", "经营持续性分数 = 经营质量水平分 × 60% + 经营质量稳定分 × 40%。"],
            ["最小历史规则", "有效财年少于5年的样本，不进入正式四类，归入观察区。"],
            ["观察区规则", "任一维度分位位于35%-65%，或经营持续性分数缺失的样本，归入观察区。分位列本身存的是0到1的小数，显示时按百分比格式呈现。"],
            ["四类定义", "长期赢家：两维分位都≥65%；审美失效/退化案例：持有分位≥65%且经营分位≤35%；经营优秀但未被长期拿住：持有分位≤35%且经营分位≥65%；弱质/陷阱样本：两维分位都≤35%。"],
            ["解释边界", "本表是研究入口，不是投资评级，也不判断机构买得对不对。"],
        ],
        columns=["方法项", "说明"],
    )
    return total, summary, method, negative


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    ws.append(df.columns.tolist())
    for row in df.replace({pd.NA: None}).itertuples(index=False, name=None):
        ws.append(list(row))
    style_sheet(ws, df)


def style_sheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    for col_name, col_idx in header_map.items():
        if col_name in PCT_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = "0.0%"
        elif col_name in SCORE_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = "0.0"
        elif col_name in DATE_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = "yyyy-mm-dd"
        elif "Debt/Equity" in str(col_name):
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = "0.00"

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).font = BOLD_FONT
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(vertical="center")

    if "分类结果" in header_map:
        class_col = header_map["分类结果"]
        fills = {
            "长期赢家": "E2F0D9",
            "审美失效/退化案例": "FCE4D6",
            "经营优秀但未被长期拿住": "DDEBF7",
            "弱质/陷阱样本": "F8CBAD",
            "观察区": "FFF2CC",
        }
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, class_col).value
            if val in fills:
                ws.cell(row, class_col).fill = PatternFill("solid", fgColor=fills[val])

    for i, column in enumerate(df.columns, start=1):
        values = [str(column)] + ["" if pd.isna(v) else str(v) for v in df[column].head(200)]
        max_len = min(max(len(v) for v in values) + 2, 30)
        if column in {"当前持有机构", "观察原因", "说明"}:
            max_len = min(max_len + 12, 60)
        ws.column_dimensions[get_column_letter(i)].width = max_len


def style_method_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).fill = SUB_FILL
        ws.cell(row, 1).font = BOLD_FONT
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 96


def main() -> None:
    total, summary, method, negative = build_output_frames()
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "分类总表", total)
    write_sheet(wb, "分类概览", summary)
    for sheet_name in ["长期赢家", "审美失效_退化案例", "经营优秀但未被长期拿住", "弱质_陷阱样本", "观察区"]:
        category = sheet_name.replace("_", "/")
        category_df = total[total["分类结果"] == category].copy()
        write_sheet(wb, sheet_name, category_df)
    write_sheet(wb, "观察与负样本参考", negative)
    ws = wb.create_sheet("方法说明")
    ws.append(method.columns.tolist())
    for row in method.itertuples(index=False, name=None):
        ws.append(list(row))
    style_method_sheet(ws)
    wb.save(OUTPUT)
    print(f"saved: {OUTPUT}")


if __name__ == "__main__":
    main()
