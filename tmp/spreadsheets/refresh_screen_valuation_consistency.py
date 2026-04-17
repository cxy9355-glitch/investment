from __future__ import annotations

import json
import math
import re
import time
from io import StringIO
from pathlib import Path
from typing import Any

import pandas as pd
import requests
import yfinance as yf
from openpyxl import load_workbook


ROOT = Path(r"G:\Codex\个人\investment")
CACHE_DIR = ROOT / "tmp" / "spreadsheets" / "cache"
SCREEN_PATH = ROOT / "机构持仓研究" / "巴芒_喜马拉雅_高瓴_历史持仓筛选_2026-04-17_可比清理版.xlsx"
FULL_DATASET_PATH = ROOT / "机构持仓研究" / "长期持有全池数据集_2026-04-17_可比清理版.xlsx"
REPORT_PATH = ROOT / "机构持仓研究" / "05_估值口径全量复核说明_2026-04-17.md"

WEB_HEADERS = {"User-Agent": "Mozilla/5.0"}
SEARCH_HEADERS = {"User-Agent": "Mozilla/5.0", "X-Requested-With": "XMLHttpRequest"}


def cached_get_text(
    url: str,
    headers: dict[str, str],
    cache_key: str,
    method: str = "GET",
    data: dict[str, Any] | None = None,
    sleep_s: float = 0.05,
) -> str:
    suffix = "post" if method.upper() == "POST" else "txt"
    cache_path = CACHE_DIR / f"{cache_key}.{suffix}"
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8")
    if method.upper() == "POST":
        resp = requests.post(url, headers=headers, data=data, timeout=40)
    else:
        resp = requests.get(url, headers=headers, timeout=40)
    resp.raise_for_status()
    cache_path.write_text(resp.text, encoding="utf-8")
    time.sleep(sleep_s)
    return resp.text


def normalize_issuer(name: str) -> str:
    cleaned = re.sub(r"[^A-Z0-9]+", " ", str(name).upper()).strip()
    stopwords = {
        "INC", "CORP", "CORPORATION", "CO", "COS", "PLC", "LTD", "LIMITED", "SA", "NV",
        "LP", "LLC", "DEL", "NEW", "HOLDINGS", "HOLDING", "GROUP", "CL", "CLASS",
        "SPONSORED", "ADR", "ADS", "ORD", "SHS", "THE",
    }
    tokens = [token for token in re.sub(r"\s+", " ", cleaned).split() if token not in stopwords]
    return " ".join(tokens)


def search_companies_marketcap(query: str) -> list[dict[str, Any]]:
    text = cached_get_text(
        "https://companiesmarketcap.com/search.do",
        headers=SEARCH_HEADERS,
        cache_key=f"cmc_search_refresh_{re.sub(r'[^A-Za-z0-9]+', '_', query)[:80]}",
        method="POST",
        data={"query": query},
    )
    return json.loads(text)


def choose_search_match(results: list[dict[str, Any]], issuer_norm: str, market: str) -> dict[str, Any] | None:
    filtered = [item for item in results if item.get("type") == "stock"]
    if not filtered:
        return None
    issuer_tokens = set(issuer_norm.split())
    best: tuple[float, dict[str, Any]] | None = None
    for item in filtered:
        name_norm = normalize_issuer(item.get("name", ""))
        name_tokens = set(name_norm.split())
        overlap = len(issuer_tokens & name_tokens) / max(1, len(issuer_tokens))
        score = overlap
        score -= 0.08 * max(0, len(name_tokens - issuer_tokens))
        score -= 0.12 * max(0, len(issuer_tokens - name_tokens))
        identifier = item.get("identifier", "")
        if "ADR" in market.upper():
            if identifier.isupper() and "." not in identifier and 1 <= len(identifier) <= 5:
                score += 0.15
        if "." in identifier or any(ch.isdigit() for ch in identifier):
            score -= 0.25
        if best is None or score > best[0]:
            best = (score, item)
    return best[1] if best else filtered[0]


def extract_history_table(html: str) -> list[float]:
    try:
        tables = pd.read_html(StringIO(html))
    except ValueError:
        return []
    for table in tables:
        cols = [str(col).strip().lower() for col in table.columns]
        if cols and "year" in cols[0] and any("ratio" in col for col in cols):
            ratio_col = next(col for col in table.columns if "ratio" in str(col).lower())
            values = pd.to_numeric(table[ratio_col], errors="coerce").dropna().tolist()
            return [float(v) for v in values if pd.notna(v)]
    return []


def extract_current_ratio(html: str, ratio_label: str) -> float | None:
    patterns = [
        rf"{ratio_label} ratio as of .*?<span class=\"background-ya\">([^<]+)</span>",
        rf"current price-to-{ratio_label.lower()} ratio .*?<strong>([^<]+)</strong>",
    ]
    for pattern in patterns:
        match = re.search(pattern, html, re.S | re.I)
        if match:
            return as_float(match.group(1).replace(",", ""))
    return None


def percentile_rank(current: float | None, history_values: list[float]) -> float | None:
    if current is None:
        return None
    valid = [v for v in history_values if v is not None and not math.isinf(v) and not math.isnan(v)]
    if not valid:
        return None
    return sum(1 for v in valid if v <= current) / len(valid)


def load_yahoo_payload(ticker: str) -> dict[str, Any]:
    cache_path = CACHE_DIR / f"yf_{ticker.replace('.', '_')}.json"
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    yf_ticker = yf.Ticker(ticker)
    payload = {"info": yf_ticker.info, "fast": dict(yf_ticker.fast_info)}
    cache_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    time.sleep(0.05)
    return payload


def fetch_usd_pair_rate(currency: str) -> float:
    if not currency or currency.upper() == "USD":
        return 1.0
    symbol = f"{currency.upper()}=X"
    payload = load_yahoo_payload(symbol)
    info = payload.get("info", {})
    fast = payload.get("fast", {})
    value = (
        fast.get("lastPrice")
        or info.get("currentPrice")
        or info.get("regularMarketPrice")
        or info.get("previousClose")
    )
    if not value:
        raise ValueError(f"missing fx rate for {currency}")
    return float(value)


def convert_amount(amount: float | None, source_currency: str | None, target_currency: str | None) -> float | None:
    if amount is None or pd.isna(amount):
        return None
    source = (source_currency or "USD").upper()
    target = (target_currency or "USD").upper()
    if source == target:
        return float(amount)
    usd_value = float(amount) / fetch_usd_pair_rate(source)
    if target == "USD":
        return usd_value
    return usd_value * fetch_usd_pair_rate(target)


def as_float(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def current_price_and_cap(payload: dict[str, Any]) -> tuple[float | None, float | None, str | None]:
    info = payload.get("info", {})
    fast = payload.get("fast", {})
    price = (
        fast.get("lastPrice")
        or info.get("currentPrice")
        or info.get("regularMarketPrice")
        or info.get("previousClose")
    )
    market_cap = info.get("marketCap")
    shares = info.get("sharesOutstanding")
    if market_cap is None and price is not None and shares is not None:
        market_cap = float(price) * float(shares)
    return (float(price) if price is not None else None, float(market_cap) if market_cap is not None else None, info.get("currency"))


def build_latest_annual() -> pd.DataFrame:
    annual = pd.read_excel(FULL_DATASET_PATH, sheet_name="年度经营质量_可比池")
    annual = annual[annual["fiscal_year"].notna()].copy()
    annual = annual.sort_values(["ticker", "fiscal_year"]).groupby("ticker", as_index=False).tail(1)
    return annual.set_index("ticker")


def is_high_risk_row(local_pe: Any, trailing_pe: Any, market: str, quote_currency: str | None, financial_currency: str | None) -> bool:
    if "ADR" in str(market).upper():
        return True
    if quote_currency and financial_currency and quote_currency != financial_currency:
        return True
    pe_local = as_float(local_pe)
    pe_yahoo = as_float(trailing_pe)
    if pe_local and pe_yahoo:
        return abs(pe_local - pe_yahoo) / abs(pe_local) >= 0.10
    return False


def refresh_rows() -> tuple[pd.DataFrame, pd.DataFrame]:
    latest_annual = build_latest_annual()
    wb = load_workbook(SCREEN_PATH)
    ws = wb["估值总表"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    idx = {name: pos + 1 for pos, name in enumerate(headers)}

    updates: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row, idx["代码"]).value
        if ticker not in latest_annual.index:
            continue
        annual_row = latest_annual.loc[ticker]
        payload = load_yahoo_payload(ticker)
        info = payload.get("info", {})
        current_price, market_cap, quote_currency = current_price_and_cap(payload)
        info_fin_currency = (info.get("financialCurrency") or quote_currency or "USD")
        fin_currency = (annual_row.get("currency") or info_fin_currency or quote_currency or "USD")
        market = ws.cell(row, idx["市场"]).value or ""
        old_pe = ws.cell(row, idx["PE(TTM)"]).value
        high_risk = is_high_risk_row(old_pe, info.get("trailingPE"), str(market), quote_currency, info_fin_currency)
        net_income = convert_amount(as_float(info.get("netIncomeToCommon")), info_fin_currency, quote_currency)
        if net_income is None or net_income <= 0:
            net_income = convert_amount(annual_row.get("net_income"), fin_currency, quote_currency)
        pe = None
        pb = ws.cell(row, idx["PB"]).value
        if market_cap and net_income and net_income > 0:
            pe = market_cap / net_income
        else:
            trailing_pe = as_float(info.get("trailingPE"))
            if trailing_pe and trailing_pe > 0:
                pe = trailing_pe
        pe_hist_pct = ws.cell(row, idx.get("PE年末历史百分位(自算)", idx.get("PE历史百分位"))).value
        pb_hist_pct = ws.cell(row, idx.get("PB年末历史百分位(自算)", idx.get("PB历史百分位"))).value
        valuation_score = ws.cell(row, idx.get("综合估值百分位(自算)", idx.get("综合估值百分位"))).value
        slug = None
        if high_risk:
            english_name = ws.cell(row, idx["英文公司名"]).value or ws.cell(row, idx["中文公司名"]).value
            match = choose_search_match(search_companies_marketcap(str(english_name)), normalize_issuer(str(english_name)), str(market))
            if match:
                slug = match.get("url")
                if slug:
                    pe_html = cached_get_text(
                        f"https://companiesmarketcap.com/{slug}/pe-ratio/",
                        headers=WEB_HEADERS,
                        cache_key=f"cmc_refresh_pe_{slug}",
                    )
                    pb_html = cached_get_text(
                        f"https://companiesmarketcap.com/{slug}/pb-ratio/",
                        headers=WEB_HEADERS,
                        cache_key=f"cmc_refresh_pb_{slug}",
                    )
                    pb = extract_current_ratio(pb_html, "P/B") or as_float(info.get("priceToBook")) or pb
                    pe_hist = [v for v in extract_history_table(pe_html) if v > 0]
                    pb_hist = [v for v in extract_history_table(pb_html) if v > 0]
                    pe_hist_pct = percentile_rank(pe, pe_hist)
                    pb_hist_pct = percentile_rank(pb, pb_hist)
                    valid = [v for v in [pe_hist_pct, pb_hist_pct] if v is not None]
                    valuation_score = sum(valid) / len(valid) if valid else None

        updates.append(
            {
                "ticker": ticker,
                "high_risk": high_risk,
                "current_price_new": current_price,
                "market_cap_100m_new": (market_cap / 1e8) if market_cap else None,
                "pe_new": pe,
                "pb_new": pb,
                "pe_hist_pct_new": pe_hist_pct,
                "pb_hist_pct_new": pb_hist_pct,
                "valuation_score_new": valuation_score,
                "quote_currency": quote_currency,
                "financial_currency": info_fin_currency,
                "cmc_slug": slug,
                "old_pe": old_pe,
                "old_pb": ws.cell(row, idx["PB"]).value,
            }
        )

    updates_df = pd.DataFrame(updates)
    return latest_annual.reset_index(), updates_df


def write_back(updates_df: pd.DataFrame) -> None:
    wb = load_workbook(SCREEN_PATH)
    valuation = wb["估值总表"]
    shortlist = wb["深研候选"]
    caliber = wb["口径说明"]

    header_renames = {
        "PE历史百分位": "PE年末历史百分位(自算)",
        "PB历史百分位": "PB年末历史百分位(自算)",
        "综合估值百分位": "综合估值百分位(自算)",
    }

    for sheet in [valuation, shortlist]:
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(1, col).value
            if value in header_renames:
                sheet.cell(1, col).value = header_renames[value]

    val_headers = [valuation.cell(1, col).value for col in range(1, valuation.max_column + 1)]
    val_idx = {name: pos + 1 for pos, name in enumerate(val_headers)}
    short_headers = [shortlist.cell(1, col).value for col in range(1, shortlist.max_column + 1)]
    short_idx = {name: pos + 1 for pos, name in enumerate(short_headers)}

    update_map = updates_df.set_index("ticker").to_dict("index")

    for sheet, idx in [(valuation, val_idx), (shortlist, short_idx)]:
        for row in range(2, sheet.max_row + 1):
            ticker = sheet.cell(row, idx["代码"]).value
            payload = update_map.get(ticker)
            if not payload:
                continue
            sheet.cell(row, idx["PE(TTM)"]).value = payload["pe_new"]
            sheet.cell(row, idx["PB"]).value = payload["pb_new"]
            sheet.cell(row, idx["当前价格"]).value = payload["current_price_new"]
            sheet.cell(row, idx["当前市值(亿美元)"]).value = payload["market_cap_100m_new"]
            if "PE年末历史百分位(自算)" in idx:
                sheet.cell(row, idx["PE年末历史百分位(自算)"]).value = payload["pe_hist_pct_new"]
            if "PB年末历史百分位(自算)" in idx:
                sheet.cell(row, idx["PB年末历史百分位(自算)"]).value = payload["pb_hist_pct_new"]
            if "综合估值百分位(自算)" in idx:
                sheet.cell(row, idx["综合估值百分位(自算)"]).value = payload["valuation_score_new"]

    obsolete_keys = {"PE(TTM)复核口径", "ADR估值修正规则", "当前PE/PB口径"}
    for row in range(caliber.max_row, 1, -1):
        if caliber.cell(row, 1).value in obsolete_keys:
            caliber.delete_rows(row, 1)

    notes = {
        "当前PE口径": "全表统一改为当前市值除以TTM净利的自算口径，跨币种先做汇率归一。",
        "当前PB口径": "PB优先沿用 CompaniesMarketCap 当前值；高风险样本已逐只刷新并与历史表保持同源。",
        "历史百分位口径": "历史百分位继续基于 CompaniesMarketCap 年末历史表自算，与富途历史分位不是同一算法。",
        "本次刷新范围": "已对全表当前PE做统一刷新，并对高风险ticker刷新PB与历史分位。",
    }
    key_to_row = {caliber.cell(row, 1).value: row for row in range(2, caliber.max_row + 1)}
    for key, value in notes.items():
        row = key_to_row.get(key)
        if row:
            caliber.cell(row, 2).value = value
        else:
            caliber.append([key, value])

    seen_keys: set[str] = set()
    for row in range(caliber.max_row, 1, -1):
        key = caliber.cell(row, 1).value
        if key in seen_keys:
            caliber.delete_rows(row, 1)
        else:
            seen_keys.add(key)

    wb.save(SCREEN_PATH)


def write_report(updates_df: pd.DataFrame) -> None:
    df = updates_df.copy()
    cross = df[df["quote_currency"].fillna("USD") != df["financial_currency"].fillna("USD")].copy()
    high_risk = df[df["high_risk"]].copy()

    lines = [
        "# 估值口径全量复核说明",
        "",
        f"- 复核工作簿：`{SCREEN_PATH.name}`",
        f"- PE统一刷新的ticker数：`{len(df)}`",
        f"- PB与历史分位全刷新的高风险ticker数：`{len(high_risk)}`",
        f"- 交易货币与财报货币不一致的ticker数：`{len(cross)}`",
        "",
        "## 复核口径",
        "",
        "- 当前PE：对全表统一改为 `当前市值 / TTM净利润` 自算；跨币种先按最新汇率换算。",
        "- 当前PB：继续优先采用 CompaniesMarketCap 当前值；高风险样本逐只刷新。",
        "- 历史百分位：继续使用 CompaniesMarketCap 年末历史表自算，只作为内部年表分位参考。",
        "",
        "## 使用建议",
        "",
        "- 当前 `PE(TTM)` 已适合继续横向比较。",
        "- 当前 `PB` 以 CompaniesMarketCap 当前值为准，更适合保留在主表中使用。",
        "- `PE/PB历史百分位(自算)` 仍不应直接拿来和富途历史分位对比。",
    ]
    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    _, updates_df = refresh_rows()
    write_back(updates_df)
    write_report(updates_df)
    print(f"saved: {SCREEN_PATH}")
    print(f"saved: {REPORT_PATH}")


if __name__ == "__main__":
    main()
